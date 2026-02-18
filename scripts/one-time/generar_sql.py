#!/usr/bin/env python3
"""
Genera el script SQL de migración a partir del archivo data.xlsx.

Hojas procesadas: TRASLADOS, RADICADOS.
Genera: scripts/migracion/migracion_datos.sql

Paso a paso:
  1. Colocar data.xlsx en la raíz del proyecto.
  2. Generar el SQL:
       python3 scripts/migracion/generar_sql.py
  3. Reemplazar el email del admin en el SQL generado:
       sed -i "s/ADMIN@EJEMPLO.COM/correo-real@dominio.com/g" scripts/migracion/migracion_datos.sql
  4. Ejecutar con psql (usar URL non-pooling, puerto 5432):
       psql "$SUPABASE_POSTGRES_URL_NON_POOLING" -f scripts/migracion/migracion_datos.sql
  5. Verificar que al final aparezca COMMIT y los conteos en NOTICE.

Notas:
  - El script usa DISABLE TRIGGER USER (no ALL) por restricciones de Supabase.
  - La columna de búsqueda del admin es 'correo' (tabla perfiles).
  - Registros sin fecha_tramite se omiten (la cuenta sí se crea).
"""

import openpyxl
from datetime import datetime, date
from collections import OrderedDict
import csv
import re
import os

# ── Configuración ──────────────────────────────────────────────
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data.xlsx')
RUNT_CSV_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'RUNT', 'organismos_transito.csv')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'migracion_datos.sql')
FECHA_MIGRACION = date.today().strftime('%Y%m%d')

# ── Helpers ────────────────────────────────────────────────────

def escape_sql(val):
    """Escapa una cadena para SQL."""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"


def format_date(val):
    """Convierte un valor de celda a formato de fecha SQL."""
    if val is None:
        return 'NULL'
    if isinstance(val, datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    if isinstance(val, date):
        return f"'{val.strftime('%Y-%m-%d')}'"
    s = str(val).strip()
    if not s or s.startswith('='):
        return 'NULL'
    # Intentar parsear
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return f"'{datetime.strptime(s, fmt).strftime('%Y-%m-%d')}'"
        except ValueError:
            continue
    return 'NULL'


def normalize_organismo(nombre):
    """Normaliza nombre de organismo para comparación."""
    return re.sub(r'\s+', ' ', nombre.strip().upper())


def extraer_municipio(nombre):
    """Intenta extraer el municipio del nombre completo del organismo."""
    n = nombre.strip().upper()
    # Patrones comunes: "... MCPAL CIUDAD", "... DE CIUDAD", "... /CIUDAD"
    # Si tiene "/" tomar lo de después
    if '/' in n:
        parts = n.split('/')
        return parts[-1].strip()
    # Patrones con ciudad al final
    patterns = [
        r'(?:MCPAL|MPAL|MUNICIPAL|DPTAL|DTAL|DISTRITAL)\s+(?:DE\s+)?(.+?)$',
        r'(?:DE|DEL)\s+(?:MUNICIPIO\s+DE\s+)?(.+?)$',
    ]
    for p in patterns:
        m = re.search(p, n)
        if m:
            municipio = m.group(1).strip()
            # Limpiar sufijos comunes
            municipio = re.sub(r'\s*/\s*\w+$', '', municipio)
            if len(municipio) > 2:
                return municipio
    return n


# ── Cargar catálogo RUNT ──────────────────────────────────────

def cargar_runt(path):
    """Carga el CSV del RUNT y construye índices para matching."""
    por_nombre = {}   # nombre_upper → {nombre, municipio, departamento}
    por_municipio = {}  # municipio_upper → [nombres]
    if not os.path.exists(path):
        print(f"⚠️  No se encontró {path}, se omite normalización RUNT.")
        return por_nombre, por_municipio
    with open(path, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            nombre = row['nombre'].strip()
            muni = row['municipio'].strip().upper()
            depto = row['departamento'].strip()
            info = {'nombre': nombre, 'municipio': muni, 'departamento': depto}
            por_nombre[nombre.upper()] = info
            por_municipio.setdefault(muni, []).append(info)
    return por_nombre, por_municipio


def simplificar(s):
    """Quita caracteres especiales y espacios extra para fuzzy match."""
    return re.sub(r'\s+', ' ', re.sub(r'[^A-Z0-9]', ' ', s.upper())).strip()


def resolver_organismo(nombre_excel, runt_nombre, runt_municipio):
    """Intenta resolver un nombre de organismo del Excel al catálogo RUNT.
    Retorna dict {nombre, municipio, departamento} o None si no hay match."""
    upper = nombre_excel.upper()

    # 1) Match exacto
    if upper in runt_nombre:
        return runt_nombre[upper]

    # 2) Match por municipio (nombres cortos como BOGOTA, CALI, etc.)
    candidatos = runt_municipio.get(upper, [])
    if not candidatos:
        # Intentar con variantes: "BOGOTA" → "BOGOTA, D.C."
        for muni_key in runt_municipio:
            if muni_key.startswith(upper + ',') or muni_key == upper:
                candidatos = runt_municipio[muni_key]
                break
    if candidatos:
        if len(candidatos) == 1:
            return candidatos[0]
        # Si hay varios, preferir el que contenga "SECRETARIA" o "STRIA"
        for c in candidatos:
            if 'SECRETARIA' in c['nombre'].upper() or 'STRIA' in c['nombre'].upper():
                return c
        return candidatos[0]

    # 3) Fuzzy: comparar versión simplificada
    s = simplificar(nombre_excel)
    for runt_key, info in runt_nombre.items():
        if simplificar(runt_key) == s:
            return info

    # 4) Match por keywords (al menos 2 palabras de 4+ chars coinciden)
    words = [w for w in upper.split() if len(w) >= 4]
    mejor, mejor_score = None, 0
    for runt_key, info in runt_nombre.items():
        score = sum(1 for w in words if w in runt_key)
        if score > mejor_score and score >= 2:
            mejor, mejor_score = info, score
    if mejor:
        return mejor

    # 5) Match por municipio parcial (ej: "PIENDOMO CAUCA" → municipio PIENDAMO)
    for muni_key, infos in runt_municipio.items():
        if upper in muni_key or muni_key in upper:
            return infos[0]
        # Levenshtein simple: 1 caracter de diferencia
        for word in upper.split():
            if len(word) >= 5 and len(muni_key) >= 5:
                if abs(len(word) - len(muni_key)) <= 1:
                    diffs = sum(1 for a, b in zip(word, muni_key) if a != b)
                    if diffs <= 1:
                        return infos[0]

    return None


print(f"Cargando catálogo RUNT desde {RUNT_CSV_PATH}...")
runt_nombre, runt_municipio = cargar_runt(RUNT_CSV_PATH)
print(f"  → {len(runt_nombre)} organismos en catálogo RUNT")

# ── Leer Excel ─────────────────────────────────────────────────

print(f"Leyendo {EXCEL_PATH}...")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# --- TRASLADOS ---
traslados = []
ws = wb['TRASLADOS']
for row in ws.iter_rows(min_row=2):
    placa = row[0].value
    if not placa:
        continue
    placa = str(placa).strip().upper()
    fecha = row[1].value
    ot_destino = str(row[2].value).strip() if row[2].value else None
    obs1 = str(row[3].value).strip() if row[3].value else None
    guia = str(row[4].value).strip() if row[4].value else None
    # Columnas 5 y 6 pueden existir o no
    fecha_guia = row[5].value if len(row) > 5 else None
    obs2 = str(row[6].value).strip() if len(row) > 6 and row[6].value else None

    # Combinar observaciones
    observaciones = None
    if obs1 and obs2:
        observaciones = f"{obs1} | {obs2}"
    elif obs1:
        observaciones = obs1
    elif obs2:
        observaciones = obs2

    # Limpiar guía (quitar .0 si es float)
    if guia and guia.endswith('.0'):
        guia = guia[:-2]

    traslados.append({
        'placa': placa,
        'fecha_tramite': fecha,
        'organismo_destino': ot_destino,
        'observaciones': observaciones,
        'numero_guia': guia,
    })

# --- RADICADOS ---
radicados = []
ws = wb['RADICADOS']
for row in ws.iter_rows(min_row=2):
    placa = row[0].value
    if not placa:
        continue
    placa = str(placa).strip().upper()
    transito = str(row[1].value).strip() if row[1].value else None
    fecha_recibo = row[2].value
    fecha_tramite = row[3].value
    # fecha_vencimiento (col 4) es fórmula, se recalculará
    obs = str(row[5].value).strip() if row[5].value else None

    radicados.append({
        'placa': placa,
        'organismo_origen': transito,
        'fecha_tramite': fecha_tramite,
        'observaciones': obs,
        'estado': 'pendiente_radicar',
    })

wb.close()

# ── Recopilar organismos únicos (normalizados con RUNT) ───────

organismos = OrderedDict()  # nombre_normalizado → {nombre, municipio, departamento}
# Mapeo: nombre_excel_normalizado → nombre_final (para usar en JOINs del SQL)
mapa_nombre = {}

sin_match_runt = []

def registrar_organismo(nombre_excel):
    """Registra un organismo, intentando resolverlo contra el RUNT."""
    nombre_norm = normalize_organismo(nombre_excel)
    if nombre_norm in mapa_nombre:
        return  # Ya procesado

    runt_info = resolver_organismo(nombre_norm, runt_nombre, runt_municipio)
    if runt_info:
        nombre_final = runt_info['nombre'].strip().upper()
        mapa_nombre[nombre_norm] = nombre_final
        if nombre_final not in organismos:
            organismos[nombre_final] = {
                'nombre': nombre_final,
                'municipio': runt_info['municipio'],
                'departamento': runt_info['departamento'],
            }
    else:
        # Sin match RUNT: usar nombre original
        sin_match_runt.append(nombre_norm)
        mapa_nombre[nombre_norm] = nombre_norm
        if nombre_norm not in organismos:
            organismos[nombre_norm] = {
                'nombre': nombre_norm,
                'municipio': extraer_municipio(nombre_norm),
                'departamento': 'POR DEFINIR',
            }

for t in traslados:
    if t['organismo_destino']:
        registrar_organismo(t['organismo_destino'])

for r in radicados:
    if r['organismo_origen']:
        registrar_organismo(r['organismo_origen'])

if sin_match_runt:
    print(f"  ⚠️  {len(sin_match_runt)} organismos sin match en RUNT:")
    for n in sin_match_runt:
        print(f"      - {n}")

# ── Recopilar placas únicas ───────────────────────────────────

todas_placas = OrderedDict()
for t in traslados:
    todas_placas[t['placa']] = True
for r in radicados:
    todas_placas[r['placa']] = True

# ── Generar SQL ────────────────────────────────────────────────

print(f"Generando SQL con {len(todas_placas)} cuentas, {len(traslados)} traslados, "
      f"{len(radicados)} radicaciones...")

lines = []
w = lines.append

w("-- ============================================================")
w("-- MIGRACIÓN DE DATOS DESDE EXCEL")
w(f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
w(f"-- Cuentas: {len(todas_placas)}")
w(f"-- Traslados: {len(traslados)}")
w(f"-- Radicaciones activas: {len(radicados)}")
w(f"-- Organismos: {len(organismos)}")
w("-- ============================================================")
w("-- INSTRUCCIONES:")
w("-- 1. Reemplazar 'ADMIN@EJEMPLO.COM' con el email del admin:")
w("--      sed -i \"s/ADMIN@EJEMPLO.COM/correo@real.com/g\" migracion_datos.sql")
w("-- 2. Ejecutar con psql (URL non-pooling, puerto 5432):")
w("--      psql \"$SUPABASE_POSTGRES_URL_NON_POOLING\" -f migracion_datos.sql")
w("-- 3. Verificar que al final aparezca COMMIT y los conteos.")
w("-- ============================================================")
w("")
w("BEGIN;")
w("")

# ── Paso 0: Variable admin ────────────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 0: Configuración")
w("-- ══════════════════════════════════════════════════════════════")
w("")
w("-- Tabla temporal con configuración de migración")
w("CREATE TEMP TABLE _mig_config (key text PRIMARY KEY, val uuid);")
w("")
w("DO $$")
w("DECLARE")
w("  v_admin uuid;")
w("  v_empresa uuid;")
w("BEGIN")
w("  -- ⚠️  CAMBIAR ESTE EMAIL POR EL DEL ADMINISTRADOR")
w("  SELECT id INTO v_admin FROM public.perfiles")
w("  WHERE correo = 'ADMIN@EJEMPLO.COM' LIMIT 1;  -- columna real en perfiles")
w("")
w("  IF v_admin IS NULL THEN")
w("    RAISE EXCEPTION 'No se encontró el usuario administrador. Verificar el email.';")
w("  END IF;")
w("")
w("  INSERT INTO _mig_config VALUES ('admin_id', v_admin);")
w("")
w("  -- Empresa transportadora Interrapidísimo")
w("  SELECT id INTO v_empresa FROM public.mov_empresas_transporte")
w("  WHERE nombre = 'INTERRAPIDISIMO' LIMIT 1;")
w("")
w("  IF v_empresa IS NOT NULL THEN")
w("    INSERT INTO _mig_config VALUES ('empresa_interrapidisimo', v_empresa);")
w("  END IF;")
w("END $$;")
w("")

# ── Paso 1: Desactivar triggers ───────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 1: Desactivar triggers problemáticos")
w("-- ══════════════════════════════════════════════════════════════")
w("")
w("-- Cuentas: desactivar todos (generaremos numero_cuenta manualmente)")
w("ALTER TABLE public.mov_cuentas_vehiculos DISABLE TRIGGER USER;")
w("")
w("-- Traslados: desactivar validaciones y historial")
w("ALTER TABLE public.mov_traslados DISABLE TRIGGER USER;")
w("")
w("-- Radicaciones: desactivar validaciones y historial")
w("ALTER TABLE public.mov_radicaciones DISABLE TRIGGER USER;")
w("")

# ── Paso 2: Insertar organismos ───────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 2: Insertar organismos de tránsito")
w("-- ══════════════════════════════════════════════════════════════")
w("")

for nombre, info in organismos.items():
    w(f"INSERT INTO public.mov_organismos_transito (nombre, tipo, municipio, departamento)")
    w(f"SELECT {escape_sql(info['nombre'])}, 'municipal', {escape_sql(info['municipio'])}, {escape_sql(info['departamento'])}")
    w(f"WHERE NOT EXISTS (SELECT 1 FROM public.mov_organismos_transito WHERE nombre = {escape_sql(info['nombre'])});")
    w("")

# ── Paso 3: Insertar cuentas ──────────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 3: Insertar cuentas de vehículos")
w("-- ══════════════════════════════════════════════════════════════")
w("")

for i, placa in enumerate(todas_placas.keys(), start=1):
    numero_cuenta = f"{FECHA_MIGRACION}-{i:05d}"
    w(f"INSERT INTO public.mov_cuentas_vehiculos (placa, numero_cuenta, tipo_servicio, creado_por)")
    w(f"SELECT {escape_sql(placa)}, {escape_sql(numero_cuenta)}, 'particular', mc.val")
    w(f"FROM _mig_config mc WHERE mc.key = 'admin_id'")
    w(f"AND NOT EXISTS (SELECT 1 FROM public.mov_cuentas_vehiculos WHERE placa = {escape_sql(placa)});")
    w("")

# ── Paso 4: Insertar traslados ────────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 4: Insertar traslados")
w("-- ══════════════════════════════════════════════════════════════")
w("")

for t in traslados:
    fecha = format_date(t['fecha_tramite'])
    nombre_norm = normalize_organismo(t['organismo_destino']) if t['organismo_destino'] else None
    organismo = escape_sql(mapa_nombre.get(nombre_norm, nombre_norm)) if nombre_norm else 'NULL'
    obs = escape_sql(t['observaciones'])
    guia = escape_sql(t['numero_guia'])

    # Si no hay organismo destino, no podemos insertar (es NOT NULL)
    if organismo == 'NULL':
        w(f"-- OMITIDO (sin organismo destino): {t['placa']}")
        w("")
        continue

    # Si no hay fecha_tramite, omitir (es NOT NULL); la cuenta ya se creó
    if fecha == 'NULL':
        w(f"-- OMITIDO (sin fecha_tramite): {t['placa']}")
        w("")
        continue

    # fecha_aprobacion = fecha_tramite (asumimos aprobación el mismo día)
    # fecha_vencimiento = sumar_dias_habiles(fecha_tramite, 60)
    w(f"INSERT INTO public.mov_traslados (")
    w(f"  cuenta_id, organismo_destino_id, estado,")
    w(f"  fecha_tramite, fecha_aprobacion, fecha_vencimiento,")
    w(f"  numero_guia, empresa_transportadora_id, observaciones,")
    w(f"  creado_por, creado_en, actualizado_en")
    w(f")")
    w(f"SELECT")
    w(f"  cv.id,")
    w(f"  ot.id,")
    w(f"  'enviado_organismo',")
    w(f"  {fecha},")
    w(f"  {fecha},")
    w(f"  CASE WHEN {fecha} IS NOT NULL THEN public.sumar_dias_habiles({fecha}::date, 60) ELSE NULL END,")
    w(f"  {guia},")
    w(f"  (SELECT val FROM _mig_config WHERE key = 'empresa_interrapidisimo'),")
    w(f"  {obs},")
    w(f"  mc.val,")
    w(f"  COALESCE({fecha}::timestamp with time zone, now()),")
    w(f"  now()")
    w(f"FROM public.mov_cuentas_vehiculos cv")
    w(f"CROSS JOIN _mig_config mc")
    w(f"JOIN public.mov_organismos_transito ot ON ot.nombre = {organismo}")
    w(f"WHERE cv.placa = {escape_sql(t['placa'])}")
    w(f"AND mc.key = 'admin_id';")
    w("")

# ── Paso 5: Insertar radicaciones activas ─────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 5: Insertar radicaciones activas (PENDIENTE RADICAR)")
w("-- ══════════════════════════════════════════════════════════════")
w("")

for r in radicados:
    fecha = format_date(r['fecha_tramite'])
    nombre_norm = normalize_organismo(r['organismo_origen']) if r['organismo_origen'] else None
    organismo = escape_sql(mapa_nombre.get(nombre_norm, nombre_norm)) if nombre_norm else 'NULL'
    obs = escape_sql(r['observaciones'])

    if organismo == 'NULL':
        w(f"-- OMITIDO (sin organismo origen): {r['placa']}")
        w("")
        continue

    # Si no hay fecha_tramite, omitir (es NOT NULL); la cuenta ya se creó
    if fecha == 'NULL':
        w(f"-- OMITIDO (sin fecha_tramite): {r['placa']}")
        w("")
        continue

    w(f"INSERT INTO public.mov_radicaciones (")
    w(f"  cuenta_id, organismo_origen_id, estado,")
    w(f"  fecha_tramite, fecha_vencimiento,")
    w(f"  observaciones, creado_por, creado_en, actualizado_en")
    w(f")")
    w(f"SELECT")
    w(f"  cv.id,")
    w(f"  ot.id,")
    w(f"  'pendiente_radicar',")
    w(f"  {fecha},")
    w(f"  CASE WHEN {fecha} IS NOT NULL THEN public.sumar_dias_habiles({fecha}::date, 60) ELSE current_date + 60 END,")
    w(f"  {obs},")
    w(f"  mc.val,")
    w(f"  COALESCE({fecha}::timestamp with time zone, now()),")
    w(f"  now()")
    w(f"FROM public.mov_cuentas_vehiculos cv")
    w(f"CROSS JOIN _mig_config mc")
    w(f"JOIN public.mov_organismos_transito ot ON ot.nombre = {organismo}")
    w(f"WHERE cv.placa = {escape_sql(r['placa'])}")
    w(f"AND mc.key = 'admin_id';")
    w("")

# ── Paso 6: Reactivar triggers ────────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 6: Reactivar triggers")
w("-- ══════════════════════════════════════════════════════════════")
w("")
w("ALTER TABLE public.mov_cuentas_vehiculos ENABLE TRIGGER USER;")
w("ALTER TABLE public.mov_traslados ENABLE TRIGGER USER;")
w("ALTER TABLE public.mov_radicaciones ENABLE TRIGGER USER;")
w("")

# ── Paso 7: Limpiar ───────────────────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 7: Limpieza")
w("-- ══════════════════════════════════════════════════════════════")
w("")
w("DROP TABLE IF EXISTS _mig_config;")
w("")

# ── Paso 8: Verificación ──────────────────────────────────────
w("-- ══════════════════════════════════════════════════════════════")
w("-- PASO 8: Verificación")
w("-- ══════════════════════════════════════════════════════════════")
w("")
w("DO $$")
w("DECLARE")
w("  v_cuentas integer;")
w("  v_traslados integer;")
w("  v_radicaciones integer;")
w("  v_organismos integer;")
w("BEGIN")
w("  SELECT count(*) INTO v_cuentas FROM public.mov_cuentas_vehiculos;")
w("  SELECT count(*) INTO v_traslados FROM public.mov_traslados;")
w("  SELECT count(*) INTO v_radicaciones FROM public.mov_radicaciones;")
w("  SELECT count(*) INTO v_organismos FROM public.mov_organismos_transito;")
w("")
w("  RAISE NOTICE '════════════════════════════════════════════';")
w("  RAISE NOTICE 'RESULTADO DE LA MIGRACIÓN:';")
w("  RAISE NOTICE '  Cuentas:       %', v_cuentas;")
w("  RAISE NOTICE '  Traslados:     %', v_traslados;")
w("  RAISE NOTICE '  Radicaciones:  %', v_radicaciones;")
w("  RAISE NOTICE '  Organismos:    %', v_organismos;")
w("  RAISE NOTICE '════════════════════════════════════════════';")
w("END $$;")
w("")
w("COMMIT;")
w("")
w("-- ══════════════════════════════════════════════════════════════")
w("-- QUERIES DE VERIFICACIÓN (ejecutar después del COMMIT)")
w("-- ══════════════════════════════════════════════════════════════")
w("")
w("-- Conteo general:")
w("-- SELECT 'cuentas' as tabla, count(*) as total FROM mov_cuentas_vehiculos")
w("-- UNION ALL SELECT 'traslados', count(*) FROM mov_traslados")
w("-- UNION ALL SELECT 'radicaciones', count(*) FROM mov_radicaciones")
w("-- UNION ALL SELECT 'organismos', count(*) FROM mov_organismos_transito;")
w("")
w("-- Traslados por estado:")
w("-- SELECT estado, count(*) FROM mov_traslados GROUP BY estado ORDER BY count(*) DESC;")
w("")
w("-- Radicaciones por estado:")
w("-- SELECT estado, count(*) FROM mov_radicaciones GROUP BY estado ORDER BY count(*) DESC;")
w("")
w("-- Procesos por vencer (radicaciones activas):")
w("-- SELECT cv.placa, r.estado, r.fecha_tramite, r.fecha_vencimiento,")
w("--   r.fecha_vencimiento - current_date as dias_restantes")
w("-- FROM mov_radicaciones r")
w("-- JOIN mov_cuentas_vehiculos cv ON cv.id = r.cuenta_id")
w("-- WHERE r.estado NOT IN ('radicado', 'devuelto')")
w("-- ORDER BY r.fecha_vencimiento;")

# ── Escribir archivo ──────────────────────────────────────────
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"\n✓ SQL generado: {OUTPUT_PATH}")
print(f"  → {len(lines)} líneas")
print(f"  → {len(todas_placas)} cuentas")
print(f"  → {len(traslados)} traslados")
print(f"  → {len(radicados)} radicaciones activas")
print(f"  → {len(organismos)} organismos")
print(f"\n⚠️  Recuerda cambiar 'ADMIN@EJEMPLO.COM' en el SQL antes de ejecutar.")
